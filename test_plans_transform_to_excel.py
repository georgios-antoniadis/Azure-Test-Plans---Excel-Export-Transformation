import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
import os
import tkinter as tk
from tkinter import filedialog
import sys

def select_csv_file():
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    file_path = filedialog.askopenfilename(
        title="Select a CSV file",
        filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
    )
    return file_path

def get_file_name(file_path):
    base_name = os.path.basename(file_path)
    file_name, _ = os.path.splitext(base_name)
    return file_name


# ------------------------------------------------------------------------------------
# STYLING
# Apply styles
header_fill = PatternFill(start_color='538DD5', end_color='538DD5', fill_type='solid')

# Title row
empty_row_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
title_row_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
alignment_center = Alignment(horizontal='center', vertical='center')

bottom_border = Border(bottom=Side(style='thin'))
# ------------------------------------------------------------------------------------

# ------------------------------------------------------------------------------------
# STAGING

# Check for command-line arguments
if len(sys.argv) > 1 and sys.argv[1] == "--test":
    # Predefined file name for testing
    predefined_file_name = "QA Sprints_Flow.csv"
    csv_file = predefined_file_name
    print(f"Test mode enabled. Using predefined file: {csv_file}")
else:
    # Default flow to select a file
    csv_file = select_csv_file()
    print(f"Selected CSV file: {csv_file}")

# Read the CSV file into a pandas DataFrame
df = pd.read_csv(csv_file)

# Remove the specified columns
columns_to_remove = ['Area Path', 'State', 'Tags']
df.drop(columns=columns_to_remove, inplace=True)

# Add the new column "Test result"
df['Test Result'] = ''
df['Comments'] = ''

# Save the DataFrame to an Excel file
file_name = get_file_name(csv_file)
output_excel_file = f"{file_name}.xlsx"
df.to_excel(output_excel_file, index=False, engine='openpyxl')
# ------------------------------------------------------------------------------------

# ------------------------------------------------------------------------------------
# PROCESSING

# Load the workbook and the active worksheet
wb = load_workbook(output_excel_file)
ws = wb.active

# Freeze the top pane
ws.freeze_panes = "A2"

# Apply the header style
for cell in ws[1]:
    cell.fill = header_fill

# Generating the title row
# Merge "ID", "Work Item Type", and "Title" columns into a single title cell
# Adding an existing entry to avoid messing up the first row
merged_rows = [(None, None, None)]

row = 2

while row < (ws.max_row + 1):
    current_id = ws[f"A{row}"].value
    current_work_item_type = ws[f"B{row}"].value
    current_title = ws[f"C{row}"].value

    if (current_id, current_work_item_type, current_title) != (None, None, None):
        print(row)
        if (current_id, current_work_item_type, current_title) not in merged_rows:
            merged_rows.append((current_id, current_work_item_type, current_title))
            
            # Add an empty row before the merged title row
            ws.insert_rows(row)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            empty_row_cell = ws.cell(row=row, column=1)
            empty_row_cell.value = None
            # empty_row_cell.fill = empty_row_fill
            empty_row_cell.alignment = alignment_center

            # Merge cells and add the title
            row += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            merged_cell = ws.cell(row=row, column=1)
            merged_cell.value = f"ID: {current_id} | Title: {current_title}"
            merged_cell.fill = title_row_fill
            merged_cell.alignment = alignment_center
    row +=1 

# Generate the final format of the excel file 
# 1. Remove unecessary columns 
columns_to_keep = ["ID", "Test Step", "Step Action", "Step Expected", "Test Result", "Comments"]
for col in list(ws.columns):
    if col[0].value not in columns_to_keep:
        ws.delete_cols(col[0].column)

# 2. Adjust column widths and enable text wrapping
column_widths = {
    "Step Action": 40,  # 4 times the default size
    "Step Expected": 40,  # 4 times the default size
    "Test Result": 20,  # 2 times the default size
    "Comments": 20  # 2 times the default size
}
for column in ws.columns:
    header = column[0].value
    if header in column_widths:
        col_letter = column[0].column_letter
        ws.column_dimensions[col_letter].width = column_widths[header]
        for cell in column:
            cell.alignment = Alignment(wrap_text=True)
# ------------------------------------------------------------------------------------

# ------------------------------------------------------------------------------------
# EXPORT

# Save to excel
wb.save(output_excel_file)

print(f"Excel file with merged titles, empty rows, adjusted column widths, and text wrapping saved as: {output_excel_file}")